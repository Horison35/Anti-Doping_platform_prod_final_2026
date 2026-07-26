#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
build_report.py (v5, исправленная) — Анализ рисков ОСФ: Рейтинг РУСАДА × прогнозная модель.

ИЗМЕНЕНИЯ ОТНОСИТЕЛЬНО v4 (по код-ревью):
  FIX-1  Дашборд: колонки «Квадрант (риск × рейтинг)» / «Зона риска (модель)» не существовали
         в итоговой таблице → KeyError. Приведено к реальным именам «Квадрант» / «Зона риска».
  FIX-2  Инвариант 5 самопроверки: отсутствие скобок вокруг сравнений при `&`
         (приоритет операторов) → TypeError. Исправлено, проверка снова работает.
  FIX-3  Добавлен обязательный лист «Не сопоставлено» (даже пустой) + вывод инварианта
         полноты с числами. Виды спорта модели без матча больше не теряются молча.
  FIX-4  Excel: цветовая заливка строк по зонам (RED/ORANGE/GREEN/NO_DATA) — требование
         промта; ширины колонок привязаны к именам, а не к индексам (не съезжают).
  FIX-5  Источник рисков: приоритетно CSV полного вывода модели (RISKS_CSV_PATH);
         парсинг графиков ipynb оставлен как fallback с громким предупреждением —
         графики содержат только топ-N видов спорта, «зелёные» вне чартов ошибочно
         получали статус «нет данных модели».
  Плюс: пути по умолчанию состыкованы с rusada_rating_watcher.py (data/ratings/osf/latest.pdf);
        plotly сделан необязательным (нет plotly → Excel строится, дашборд пропускается
        с предупреждением, код выхода 0); динамический диапазон оси proba.

ЭКСПОРТ ИЗ НОУТБУКА МОДЕЛИ (одна строка, добавить в конец прогона predict):
  итог_прогноза[["Вид спорта","зона_риска","proba","причина"]].to_csv("model_risks.csv", index=False)

ЗАПУСК:
  python build_report.py                         # пути по умолчанию
  python build_report.py --pdf path.pdf --risks-csv model_risks.csv --out ./reports
"""

import argparse
import json
import re
import sys
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Правила приоритизации живут только в siar/rules.py (STRUCTURE.md, п. 3) —
# этот модуль их больше не дублирует, только вызывает evaluate().
ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))
from siar.rules import (  # noqa: E402
    EntityKind, RATING_SCALES, evaluate, is_criterion_met,
)
from siar.xlsx_determinism import normalize_xlsx_timestamps  # noqa: E402

try:
    import pdfplumber
except ImportError:
    print("❌ Требуется pdfplumber:  pip install pdfplumber")
    sys.exit(2)

try:
    import plotly.graph_objects as go
    import plotly.express as px
    HAS_PLOTLY = True
except ImportError:
    HAS_PLOTLY = False

BASE_DIR = Path(__file__).resolve().parent

# ── ПУТИ ПО УМОЛЧАНИЮ (состыкованы с rusada_rating_watcher.py) ─────────────
DEFAULT_PDF   = BASE_DIR / "data" / "ratings" / "osf" / "latest.pdf"
DEFAULT_CSV   = BASE_DIR / "model_risks.csv"          # полный вывод модели (приоритетный источник)
DEFAULT_IPYNB = BASE_DIR / "2_predict.ipynb"          # fallback: графики ноутбука (неполно!)
DEFAULT_OUT   = BASE_DIR

# ── СТИЛЬ ──────────────────────────────────────────────────────────────────
INK, SUB, GRID = "#0F2D52", "#7C8DA6", "#EEF2F7"
FONT = "Inter, Segoe UI, Roboto, Arial, sans-serif"

ZONE_CODE_RED, ZONE_CODE_ORANGE = "RED", "ORANGE"
ZONE_CODE_GREEN, ZONE_CODE_NODATA = "GREEN", "NO_DATA"

ZONE_EMOJI  = {"RED": "🔴", "ORANGE": "🟠", "GREEN": "🟢", "NO_DATA": "⚪"}
ZONE_COLORS = {"RED": "#DC2626", "ORANGE": "#F59E0B", "GREEN": "#10B981", "NO_DATA": "#94A3B8"}
# FIX-4: заливка строк Excel по зонам (мягкие оттенки, текст остаётся читаемым)
ZONE_FILLS  = {"RED": "FDE2E2", "ORANGE": "FDF0D5", "GREEN": "DEF7EC", "NO_DATA": "F1F5F9"}

QUADRANT_COLORS = {
    "Приоритет 1 (низкий рейтинг + систематичность)":      "#DC2626",
    "Приоритет 2 (высокий рейтинг + систематичность)":     "#F59E0B",
    "Приоритет 3 (низкий рейтинг + нет систематичности)":  "#3B82F6",
    "Приоритет 4 (высокий рейтинг + нет систематичности)": "#10B981",
}
# Правило: scatter — по QUADRANT_COLORS, бар и Excel — по ZONE_COLORS. Не смешивать.
# Текст квадранта — отображение уровня отчёта (не формулировка LOGIC.md §3,
# rules.evaluate() отдаёт только priority); те же 4 строки, что ключи выше.
QUADRANTS = {i + 1: label for i, label in enumerate(QUADRANT_COLORS)}

CRITERIA = ["Стратегия", "План-график", "Регионы", "Сайт",
            "Семинар", "Соглашение", "Допуск", "Мониторинг", "Инфо"]
# Порог не хардкодим — берём из siar/rules.py (единственное место правил).
RATING_THRESHOLD = RATING_SCALES[EntityKind.OSF].high_threshold

ZONE_ORDER = {"🔴 КРАСНАЯ": 0, "🟠 ОРАНЖЕВАЯ": 1, "🟢 ЗЕЛЁНАЯ": 2,
              "RED": 0, "ORANGE": 1, "GREEN": 2, "NO_DATA": 3}


# ══════════════════════════ НОРМАЛИЗАЦИЯ И МАТЧИНГ ══════════════════════════

STOPWORDS = {
    "федерация", "всероссийская", "всероссийское", "российская", "российский",
    "союз", "ассоциация", "объединение", "объединенная", "национальная",
    "национальный", "общероссийская", "спортивная", "спортивно", "спорта",
    "спорт", "россии", "россия", "лиц", "с", "и", "на", "в", "по", "совет",
    "офсоо", "федерации", "видов", "года", "деятельности", "антидопинговой",
    "кинологической", "системе", "заболеванием",
}

ALIASES = {
    "плавание": "водных видов спорта", "прыжки в воду": "водных видов спорта",
    "водное поло": "водных видов спорта", "синхронное плавание": "водных видов спорта",
    "легкая атлетика": "легкой атлетики",
    "велоспорт": "велосипедного спорта", "велоспорт шоссе": "велосипедного спорта",
    "велоспорт трек": "велосипедного спорта", "маунтинбайк": "велосипедного спорта",
    "бмх": "велосипедного спорта",
    "гребля": "гребного спорта", "гребной спорт": "гребного спорта",
    "академическая гребля": "гребного спорта",
    "гребля на байдарках и каноэ": "каноэ", "гребной слалом": "каноэ",
    "тяжелая атлетика": "тяжелой атлетики",
    "конькобежный спорт": "конькобежцев", "коньки": "конькобежцев", "шорт-трек": "конькобежцев",
    "фигурное катание": "фигурного катания",
    "лыжные гонки": "лыжных гонок",
    "лыжное двоеборье": "прыжков на лыжах", "прыжки на лыжах с трамплина": "прыжков на лыжах",
    "биатлон": "биатлонистов",
    "футбол": "футбольный союз", "мини-футбол": "футбольный союз",
    "баскетбол": "баскетбола", "гандбол": "гандбола",
    "волейбол": "волейбола", "пляжный волейбол": "волейбола",
    "борьба": "спортивной борьбы", "вольная борьба": "спортивной борьбы",
    "греко-римская борьба": "спортивной борьбы",
    "самбо": "самбо", "дзюдо": "дзюдо", "бокс": "бокса", "кикбоксинг": "кикбоксинга",
    "тхэквондо": "тхэквондо", "каратэ": "каратэ",
    "муайтай": "муайтай", "тайский бокс": "муайтай",
    "пауэрлифтинг": "пауэрлифтинга", "бодибилдинг": "бодибилдинга",
    "художественная гимнастика": "гимнастики", "спортивная гимнастика": "гимнастики",
    "прыжки на батуте": "гимнастики",
    "теннис": "тенниса", "настольный теннис": "настольного тенниса",
    "гольф": "гольфа", "регби": "регби",
    "хоккей": "хоккея россии", "хоккей на траве": "хоккея на траве",
    "хоккей с мячом": "хоккея с мячом",
    "триатлон": "триатлона", "современное пятиборье": "современного пятиборья",
    "стрельба из лука": "стрельбы из лука",
    "пулевая стрельба": "пулевой стрельбы", "стендовая стрельба": "пулевой стрельбы",
    "фехтование": "фехтования", "конный спорт": "конного спорта",
    "бадминтон": "бадминтона", "скалолазание": "скалолазания",
    "сноуборд": "сноуборда", "фристайл": "фристайла",
    "прыжки в воду вышка": "водных видов спорта",
    "велосипедный спорт": "велосипедного спорта",
    "водные виды": "водных видов спорта",
    "водные поло": "водных видов спорта",
    "федерация водного поло россии": "водных видов спорта",
    "хоккей с шайбой": "хоккея россии",
    "пауэрлифтинг бодибилдинг": "бодибилдинга",
    "пауэрлифтинг тяжелая атлетика": "тяжелой атлетики",
    "горнолыжный": "горнолыжного спорта",
    "горные лыжи сноуборд": "горнолыжного спорта",
    "сноубординг": "сноуборда",
    "армрестлинг армспорт": "армрестлинга",
    "бальные танцы": "танцевального спорта",
    "мотоспорт": "мотоциклетного спорта",
    "тхэквондо втф": "союз тхэквондо",
}


def _norm(s):
    s = str(s).lower().replace("ё", "е")
    s = s.replace("«", " ").replace("»", " ").replace('"', " ").replace("'", " ")
    s = re.sub(r"[\-–—]", " ", s)
    s = re.sub(r"\([^)]*\)", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def _sport_tokens(name):
    return {w for w in _norm(name).split() if w not in STOPWORDS and len(w) > 2}


def _tok_eq(a, b):
    """Нечёткое равенство токенов, устойчивое к падежам: «плавание»≈«плавания»,
    «бокса»≈«бокс», но «бокс»≠«кикбоксинг» (общий префикс должен покрывать
    почти весь более длинный токен). Подстрочные вхождения по-прежнему запрещены."""
    if a == b:
        return True
    common = 0
    for ca, cb in zip(a, b):
        if ca != cb:
            break
        common += 1
    shortest = max(4, min(len(a), len(b)) - 2)
    return common >= shortest


def _covers(s_tokens, o_tokens):
    """Каждый значимый токен модели находит нечётко-равный токен в названии ОСФ."""
    return all(any(_tok_eq(st, ot) for ot in o_tokens) for st in s_tokens)


PARA_MARKERS = [
    ("пода", "лиц с пода"),
    ("слепых", "спорта слепых"),
    ("глухих", "спорта глухих"),
    ("сурдлимпийский", "спорта глухих"),
    ("лин ", "спорта лин"),
    (" лин", "спорта лин"),
]


def match_sport_to_osf(sport_name, osf_df):
    sn = _norm(sport_name)
    s_tokens = _sport_tokens(sport_name)

    sn_raw = str(sport_name).lower().replace("ё", "е")
    for marker, target in PARA_MARKERS:
        if marker in f" {sn} " or marker in f" {sn_raw} ":
            for _, row in osf_df.iterrows():
                if target in row["_key"]:
                    return row, "пара-профиль", 0.9

    if sn in ALIASES:
        target = _norm(ALIASES[sn])
        for _, row in osf_df.iterrows():
            if target in row["_key"]:
                return row, "алиас", 1.0

    best, best_score = None, 0.0
    for _, row in osf_df.iterrows():
        o_tokens = row["_tokens"]
        if not s_tokens or not o_tokens:
            continue
        if _covers(s_tokens, o_tokens):
            # предпочитаем ОСФ с минимумом «лишних» токенов (точнее соответствие)
            score = len(s_tokens) / max(len(o_tokens), 1)
            if score > best_score:
                best, best_score = row, score
    if best is not None:
        return best, "токены", round(best_score, 2)

    best, best_hits = None, 0
    for _, row in osf_df.iterrows():
        hits = sum(1 for st in s_tokens if any(_tok_eq(st, ot) for ot in row["_tokens"]))
        if hits > best_hits:
            best, best_hits = row, hits
    if best is not None and (best_hits >= 2 or (best_hits == 1 and len(s_tokens) == 1)):
        return best, "частичный токен", round(best_hits / max(len(s_tokens), 1), 2)

    return None, "нет матча", 0.0


# ══════════════════════════ ШАГ 1. РЕЙТИНГ ИЗ PDF ══════════════════════════

def load_osf_rating(pdf_path):
    raw = []
    with pdfplumber.open(pdf_path) as pdf:
        n_pages = len(pdf.pages)
        for page in pdf.pages:
            for table in page.extract_tables():
                raw.extend(table)
    rows = []
    for r in raw:
        if not r or r[0] is None:
            continue
        name = re.sub(r"\s+", " ", str(r[0]).replace("\n", " ")).strip()
        if name in ("", "Федерация"):
            continue
        vals = []
        for v in r[1:10]:
            try:
                vals.append(int(str(v).strip()))
            except Exception:
                vals.append(0)
        try:
            score = int(str(r[10]).strip())
        except Exception:
            score = None
        try:
            place = int(str(r[11]).strip())
        except Exception:
            place = None
        if name and score is not None:
            rows.append([name] + vals + [score, place])

    df = pd.DataFrame(rows, columns=["ОСФ"] + CRITERIA + ["Баллы", "Место"])
    df = df.drop_duplicates(subset=["ОСФ"]).reset_index(drop=True)
    df["Выполненные критерии"]   = df.apply(lambda r: "; ".join(c for c in CRITERIA if r[c] >= 1), axis=1)
    df["Невыполненные критерии"] = df.apply(lambda r: "; ".join(c for c in CRITERIA if r[c] < 1), axis=1)
    df["_key"] = df["ОСФ"].apply(_norm)
    df["_tokens"] = df["ОСФ"].apply(_sport_tokens)
    print(f"   страниц PDF: {n_pages} · извлечено сырых строк: {len(raw)} · федераций после очистки: {len(df)}")
    return df


# ══════════════════════════ ШАГ 2. РИСКИ МОДЕЛИ ══════════════════════════

def load_risks_from_csv(csv_path):
    """FIX-5: приоритетный источник — полный табличный вывод модели."""
    df = pd.read_csv(csv_path)
    need = {"Вид спорта", "зона_риска", "proba"}
    missing = need - set(df.columns)
    if missing:
        raise ValueError(f"в {csv_path} нет колонок: {missing}")
    if "причина" not in df.columns:
        df["причина"] = ""
    df["_z"] = df["зона_риска"].map(ZONE_ORDER).fillna(3)
    df = df.sort_values(["_z", "proba"], ascending=[True, False]).drop(columns="_z")
    agg = df.groupby("Вид спорта", as_index=False).first()
    agg["_key"] = agg["Вид спорта"].apply(_norm)
    agg["_tokens"] = agg["Вид спорта"].apply(_sport_tokens)
    return agg


def load_risks_from_ipynb(ipynb_path):
    """Fallback: парсинг сохранённых plotly-выводов ноутбука.
    ВНИМАНИЕ: графики содержат только топ-N видов спорта — виды вне чартов
    получат статус «нет данных модели», что искажает приоритеты 3/4.
    Использовать только пока модель не отдаёт CSV."""
    with open(ipynb_path, "r", encoding="utf-8") as f:
        nb = json.load(f)
    records = []
    for cell in nb.get("cells", []):
        for output in cell.get("outputs", []):
            pdata = output.get("data", {}).get("application/vnd.plotly.v1+json", {})
            if not pdata:
                continue
            for trace in pdata.get("data", []):
                y_vals, cdata, txt = trace.get("y", []), trace.get("customdata", []), trace.get("text", [])
                if not y_vals or not cdata:
                    continue
                for i, label in enumerate(y_vals):
                    try:
                        records.append({
                            "Вид спорта": str(label).split(" — ")[0].strip(),
                            "зона_риска": (cdata[i][1] if i < len(cdata) and len(cdata[i]) > 1 else "🟢 ЗЕЛЁНАЯ"),
                            "proba": float(txt[i]) if i < len(txt) else 0.0,
                            "причина": (cdata[i][0] if i < len(cdata) and len(cdata[i]) > 0 else ""),
                        })
                    except Exception:
                        continue
    df = pd.DataFrame(records)
    if df.empty:
        return df
    df["_z"] = df["зона_риска"].map(ZONE_ORDER).fillna(3)
    df = df.sort_values(["_z", "proba"], ascending=[True, False]).drop(columns="_z")
    agg = df.groupby("Вид спорта", as_index=False).first()
    agg["_key"] = agg["Вид спорта"].apply(_norm)
    agg["_tokens"] = agg["Вид спорта"].apply(_sport_tokens)
    return agg


# ══════════════════════════ ШАГ 3–4. СБОРКА ТАБЛИЦЫ ══════════════════════════

def build_final_table(osf_df, risks_df):
    osf_to_risk, audit, unmatched = {}, [], []   # FIX-3: копим несопоставленное

    if not risks_df.empty:
        for _, risk in risks_df.iterrows():
            osf_row, mtype, conf = match_sport_to_osf(risk["Вид спорта"], osf_df)
            audit.append({
                "Вид спорта (модель)": risk["Вид спорта"],
                "ОСФ (рейтинг)": osf_row["ОСФ"] if osf_row is not None else "—",
                "Тип матчинга": mtype, "Уверенность": conf,
                "Зона": risk["зона_риска"], "proba": round(float(risk["proba"]), 3),
            })
            if osf_row is None:
                unmatched.append({
                    "Название": risk["Вид спорта"], "Источник": "модель",
                    "Причина": "ни алиас, ни токены не дали соответствия в рейтинге ОСФ",
                })
                continue
            osf_name = osf_row["ОСФ"]
            prev = osf_to_risk.get(osf_name)
            z_new = ZONE_ORDER.get(risk["зона_риска"], 3)
            if prev is None or z_new < ZONE_ORDER.get(prev["зона_риска"], 3) or (
                z_new == ZONE_ORDER.get(prev["зона_риска"], 3)
                and float(risk["proba"]) > float(prev["proba"])
            ):
                osf_to_risk[osf_name] = {
                    "зона_риска": risk["зона_риска"], "proba": float(risk["proba"]),
                    "причина": risk["причина"], "вид_модель": risk["Вид спорта"],
                }

    results = []
    for _, osf in osf_df.iterrows():
        rating = int(osf["Баллы"])
        not_done_c = osf["Невыполненные критерии"]
        done_c = osf["Выполненные критерии"]
        unmet = [c for c in CRITERIA if not is_criterion_met(osf[c])]
        risk = osf_to_risk.get(osf["ОСФ"])
        miss_str = not_done_c if not_done_c else "нет"
        done_str = done_c if done_c else "нет"

        if risk is None:
            zone_code, zone, proba, reason = ZONE_CODE_NODATA, "⚪ NO_DATA", np.nan, None
        else:
            zr = risk["зона_риска"]
            if zr in ("🔴 КРАСНАЯ", "RED"):
                zone_code, zone = ZONE_CODE_RED, "🔴 RED"
            elif zr in ("🟠 ОРАНЖЕВАЯ", "ORANGE"):
                zone_code, zone = ZONE_CODE_ORANGE, "🟠 ORANGE"
            elif zr in ("🟢 ЗЕЛЁНАЯ", "GREEN"):
                zone_code, zone = ZONE_CODE_GREEN, "🟢 GREEN"
            else:
                zone_code, zone = ZONE_CODE_NODATA, "⚪ NO_DATA"
            proba = risk["proba"]
            reason = str(risk["причина"]) or None

        # Приоритет/состояние/обоснование/рекомендация — только через rules.evaluate()
        # (LOGIC.md §3, дословные формулировки; никакой генерации «на лету» здесь).
        try:
            assessment = evaluate(EntityKind.OSF, zone_code, rating, reason=reason,
                                  unmet_criteria=unmet, object_name=osf["ОСФ"])
        except ValueError as e:
            sys.exit(f"❌ SIAR v2 отклонил «{osf['ОСФ']}»: {e}")

        priority, quadrant = assessment.priority, QUADRANTS[assessment.priority]
        rec, justification = assessment.recommendation, assessment.justification

        results.append({
            "Вид спорта (ОСФ)": osf["ОСФ"], "Приоритет": priority, "Квадрант": quadrant,
            "Зона риска": zone, "_zone_code": zone_code,
            "Оценка риска (proba)": round(proba, 3) if proba == proba else None,
            "Баллы рейтинга РУСАДА": rating,
            "Место в рейтинге": osf["Место"] if pd.notna(osf["Место"]) else "",
            "Выполненные критерии РУСАДА": done_str,
            "Невыполненные критерии РУСАДА": miss_str,
            "Рекомендация": rec, "Обоснование": justification,
        })

    df_out = pd.DataFrame(results)
    zone_sort = {ZONE_CODE_RED: 1, ZONE_CODE_ORANGE: 2, ZONE_CODE_GREEN: 3, ZONE_CODE_NODATA: 4}
    df_out["_zone_sort"] = df_out["_zone_code"].map(zone_sort).fillna(4)
    df_out["_p"] = df_out["Оценка риска (proba)"].fillna(-1)
    df_out = (df_out.sort_values(["Приоритет", "_zone_sort", "_p"], ascending=[True, True, False])
                    .drop(columns=["_zone_sort", "_p"]).reset_index(drop=True))
    df_out.insert(0, "№", range(1, len(df_out) + 1))

    audit_df = pd.DataFrame(audit)
    unmatched_df = pd.DataFrame(unmatched, columns=["Название", "Источник", "Причина"])

    # ── САМОПРОВЕРКА ──
    ok = True
    n_rating, n_matrix = len(osf_df), len(df_out)
    print(f"\n{'='*64}\nСАМОПРОВЕРКА (инварианты)\n{'='*64}")
    c1 = n_matrix == n_rating
    ok &= c1
    print(f"1. Строк в итоге ({n_matrix}) = N_рейтинг ({n_rating}): {'✓' if c1 else '✗ ОШИБКА'}")
    print(f"   Несопоставленных со стороны модели: {len(unmatched_df)} (в лист «Не сопоставлено»)")
    prio_dist = df_out['Приоритет'].value_counts().sort_index().to_dict()
    c2 = sum(prio_dist.values()) == n_matrix
    ok &= c2
    print(f"2. Сумма по приоритетам {prio_dist} = {sum(prio_dist.values())}: {'✓' if c2 else '✗ ОШИБКА'}")
    zone_dist = df_out['_zone_code'].value_counts().to_dict()
    print(f"3. Распределение по зонам: {zone_dist}")
    bad4 = df_out[df_out["_zone_code"].isin([ZONE_CODE_RED, ZONE_CODE_ORANGE]) & df_out["Приоритет"].isin([3, 4])]
    c4 = len(bad4) == 0
    ok &= c4
    print(f"4. RED/ORANGE в приоритетах 3–4: {len(bad4)}: {'✓' if c4 else '✗ ОШИБКА'}")
    # FIX-2: скобки — теперь проверка реально выполняется
    bad5 = df_out[(df_out["_zone_code"] == ZONE_CODE_GREEN) & df_out["Приоритет"].isin([1, 2])]
    c5 = len(bad5) == 0
    ok &= c5
    print(f"5. GREEN в приоритетах 1–2: {len(bad5)}: {'✓' if c5 else '✗ ОШИБКА'}")
    hi = df_out["Баллы рейтинга РУСАДА"] >= RATING_THRESHOLD
    bad6 = df_out[(hi & df_out["Приоритет"].isin([1, 3])) | (~hi & df_out["Приоритет"].isin([2, 4]))]
    c6 = len(bad6) == 0
    ok &= c6
    print(f"6. Рейтинг ≥{RATING_THRESHOLD} в приоритетах 1/3 или <{RATING_THRESHOLD} в 2/4: {len(bad6)}: {'✓' if c6 else '✗ ОШИБКА'}")
    print(f"{'='*64}: {'ВСЕ ИНВАРИАНТЫ ПРОЙДЕНЫ' if ok else 'ЕСТЬ НАРУШЕНИЯ — ВЫДАЧА ЗАПРЕЩЕНА'}\n")
    if not ok:
        sys.exit(3)
    return df_out, audit_df, unmatched_df


# ══════════════════════════ ШАГ 5. EXCEL ══════════════════════════

COL_WIDTHS = {  # FIX-4: по именам, не по индексам
    "№": 5, "Вид спорта (ОСФ)": 45, "Приоритет": 10, "Квадрант": 30,
    "Зона риска": 14, "Оценка риска (proba)": 13, "Баллы рейтинга РУСАДА": 12,
    "Место в рейтинге": 10, "Выполненные критерии РУСАДА": 38,
    "Невыполненные критерии РУСАДА": 38, "Рекомендация": 52, "Обоснование": 62,
}


def save_excel(df, audit_df, unmatched_df, path):
    export_cols = [c for c in df.columns if not c.startswith("_")]
    df_export = df[export_cols].copy()
    zone_codes = df["_zone_code"].tolist()

    with pd.ExcelWriter(path, engine="openpyxl") as xl:
        df_export.to_excel(xl, index=False, sheet_name="Итог")
        # FIX-3: лист обязателен, даже пустой
        unmatched_df.to_excel(xl, index=False, sheet_name="Не сопоставлено")
        if audit_df is not None and not audit_df.empty:
            audit_df.to_excel(xl, index=False, sheet_name="Аудит матчинга")

    wb = openpyxl.load_workbook(path)
    ws = wb["Итог"]

    header_fill = PatternFill("solid", fgColor="0F2D52")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    for cell in ws[1]:
        cell.fill, cell.font = header_fill, header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for idx, col_name in enumerate(df_export.columns, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = COL_WIDTHS.get(col_name, 18)

    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row)):
        # FIX-4: заливка всей строки по зоне
        fill = PatternFill("solid", fgColor=ZONE_FILLS.get(zone_codes[i], "FFFFFF"))
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            cell.fill = fill

    ws.row_dimensions[1].height = 42
    for i in range(2, ws.max_row + 1):
        ws.row_dimensions[i].height = 52
    ws.freeze_panes = "B2"
    # Воспроизводимость (LOGIC.md §1): openpyxl иначе пишет created/modified =
    # текущее время в docProps/core.xml — при идентичном входе получался бы
    # неидентичный файл байт-в-байт. Данные это не меняет, только метаданные.
    _FIXED_TIMESTAMP = datetime(2000, 1, 1)
    wb.properties.created = _FIXED_TIMESTAMP
    wb.properties.modified = _FIXED_TIMESTAMP
    wb.save(path)
    normalize_xlsx_timestamps(path)  # zip-контейнер: см. siar/xlsx_determinism.py
    print(f"✅ Excel сохранён: {path}")


# ══════════════════════════ ШАГ 6. ДАШБОРД ══════════════════════════

# Приведение обрезанного названия ОСФ к читаемой подписи в именительном падеже.
_LABEL_FIX = {
    "спортивной борьбы": "Спортивная борьба", "тяжелой атлетики": "Тяжёлая атлетика",
    "тяжёлой атлетики": "Тяжёлая атлетика", "легкой атлетики": "Лёгкая атлетика",
    "водных видов спорта": "Водные виды спорта", "велосипедного спорта": "Велоспорт",
    "конькобежцев": "Конькобежный спорт", "биатлонистов": "Биатлон",
    "гребного спорта": "Гребной спорт", "фигурного катания на коньках": "Фигурное катание",
    "лыжных гонок": "Лыжные гонки", "художественной гимнастики": "Художественная гимнастика",
    "гимнастики": "Гимнастика", "тенниса": "Теннис", "настольного тенниса": "Настольный теннис",
    "бокса": "Бокс", "дзюдо": "Дзюдо", "самбо": "Самбо", "регби": "Регби",
    "хоккея": "Хоккей", "гандбола": "Гандбол", "волейбола": "Волейбол", "баскетбола": "Баскетбол",
    "футбольный союз": "Футбол", "плавания": "Плавание", "бадминтона": "Бадминтон",
    "тхэквондо": "Тхэквондо", "каратэ": "Каратэ", "бодибилдинга": "Бодибилдинг",
    "пауэрлифтинга": "Пауэрлифтинг", "армрестлинга": "Армрестлинг", "фехтования": "Фехтование",
    "сноуборда": "Сноуборд", "фристайла": "Фристайл", "скалолазания": "Скалолазание",
    "триатлона": "Триатлон", "гольфа": "Гольф", "каноэ": "Гребля на каноэ",
    "конного спорта": "Конный спорт", "стрельбы из лука": "Стрельба из лука",
    "керлинга": "Кёрлинг", "кёрлинга": "Кёрлинг", "гребного спорта": "Гребной спорт",
    "спортивной борьбы": "Спортивная борьба", "прыжков на лыжах": "Прыжки на лыжах",
    "спорта слепых": "Спорт слепых", "спорта глухих": "Спорт глухих",
    "лиц с пода": "Спорт лиц с ПОДА", "спорта лин": "Спорт ЛИН",
    "бейсбола и софтбола": "Бейсбол и софтбол", "подводного спорта": "Подводный спорт",
    "гиревого спорта": "Гиревой спорт", "бобслея": "Бобслей", "сумо": "Сумо",
    "кикбоксинга": "Кикбоксинг", "муайтай": "Муай-тай", "го": "Го", "сквоша": "Сквош",
}


def _short_name(s):
    s = str(s).replace("\n", " ").strip()
    core = s
    for prefix in ["Всероссийская федерация ", "Всероссийское спортивно-кинологическое объединение ",
                   "Общероссийская спортивная федерация ", "Объединенная федерация ",
                   "Национальная федерация ", "Российская Федерация ", "Российская федерация ",
                   "Федерация ", "Российский ", "Национальный ", "Спортивная Федерация ",
                   "Союз ", "Ассоциация "]:
        if core.startswith(prefix):
            core = core[len(prefix):]
            break
    core = core.replace(" России", "").replace(" России", "").strip()
    key = core.lower().strip(' "«»')
    if key in _LABEL_FIX:
        return _LABEL_FIX[key]
    # иначе — первая буква заглавная, остальное как есть
    return core[:1].upper() + core[1:] if core else s


def _theme(fig, title, subtitle=""):
    sub_html = f"<br><span style='font-size:12px;color:{SUB}'>{subtitle}</span>" if subtitle else ""
    fig.update_layout(template="plotly_white", font=dict(family=FONT, color=INK, size=13),
                      title=dict(text=f"<b>{title}</b>{sub_html}", x=0.01, xanchor="left", y=0.97,
                                 font=dict(size=18)),
                      margin=dict(l=10, r=24, t=84, b=24),
                      paper_bgcolor="white", plot_bgcolor="white", legend_title_text="")
    fig.update_xaxes(showgrid=False, zeroline=False)
    fig.update_yaxes(showgrid=False, zeroline=False)
    return fig


def build_dashboard(df, path):
    if not HAS_PLOTLY:
        print("⚠ plotly не установлен — дашборд пропущен (pip install plotly). Excel готов.")
        return
    df_all = df.copy()
    df_all["_short"] = df_all["Вид спорта (ОСФ)"].apply(_short_name)
    df_risk = df_all[df_all["_zone_code"].isin([ZONE_CODE_RED, ZONE_CODE_ORANGE])].copy()

    # Динамический потолок оси proba: реальные значения малы, [0,1] сплющит картину
    p_max = float(df_all["Оценка риска (proba)"].max() or 0)
    y_top = max(0.1, p_max * 1.25)

    # ── ГРАФИК 1: матрица квадрантов ── (FIX-1: реальные имена колонок)
    fig2 = go.Figure()
    rng = np.random.default_rng(42)
    for quad in QUADRANT_COLORS:
        sub = df_all[df_all["Квадрант"] == quad].copy()
        if sub.empty:
            continue
        y = sub["Оценка риска (proba)"].astype(float).fillna(-y_top * 0.07).values
        x = sub["Баллы рейтинга РУСАДА"].astype(float).values
        x = x + rng.uniform(-2.0, 2.0, size=len(x))
        y = y + rng.uniform(-y_top * 0.015, y_top * 0.015, size=len(y))
        fig2.add_trace(go.Scatter(
            x=x, y=y, mode="markers", name=quad,
            marker=dict(color=QUADRANT_COLORS[quad], size=12, opacity=0.82,
                        line=dict(width=1, color="white")),
            customdata=np.stack([sub["Вид спорта (ОСФ)"].values,
                                 sub["Зона риска"].values,
                                 sub["Обоснование"].values], axis=-1),
            hovertemplate=("<b>%{customdata[0]}</b><br>Рейтинг: %{x:.0f} баллов<br>"
                           "Зона: %{customdata[1]}<br><i>%{customdata[2]}</i><extra></extra>"),
        ))
    fig2.add_vline(x=RATING_THRESHOLD, line_dash="dot", line_color="#94A3B8", line_width=1.5,
                   annotation_text=f"порог рейтинга = {RATING_THRESHOLD}",
                   annotation_position="top left", annotation_font=dict(color="#94A3B8", size=10))
    fig2.add_hline(y=0.0, line_dash="dot", line_color="#CBD5E1", line_width=1)
    _theme(fig2, "Матрица Риск × Рейтинг ОСФ 2025",
           "ось X — баллы рейтинга РУСАДА · ось Y — оценка риска модели (⚪ нет данных ≈ ниже 0)")
    fig2.update_layout(height=700,
                       xaxis=dict(range=[0, 110], title="Баллы рейтинга РУСАДА", showgrid=True, gridcolor=GRID),
                       yaxis=dict(range=[-y_top * 0.14, y_top], title="Оценка риска (proba)",
                                  showgrid=True, gridcolor=GRID),
                       legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1))

    # ── ГРАФИК 2: топ рисковых ──
    if not df_risk.empty:
        d1 = df_risk.sort_values("Оценка риска (proba)").tail(20).copy()
        fig1 = px.bar(d1, x="Оценка риска (proba)", y="_short", orientation="h",
                      color="_zone_code", color_discrete_map=ZONE_COLORS,
                      text=d1["Оценка риска (proba)"].map(lambda v: f"{v:.2f}"),
                      hover_name="Вид спорта (ОСФ)",
                      hover_data={"Баллы рейтинга РУСАДА": True, "Квадрант": True, "_short": False})
        fig1.update_traces(marker_cornerradius=7, textposition="outside",
                           cliponaxis=False, textfont=dict(size=11, color=INK))
        _theme(fig1, "Рисковые виды спорта (ОСФ)",
               "цвет = зона: 🔴 свежая динамика · 🟠 история · топ-20 по оценке модели")
        fig1.update_layout(height=max(420, 32 * len(d1)), yaxis_title="",
                           xaxis_title="оценка риска", legend_title_text="зона")
        fig1.update_xaxes(showgrid=True, gridcolor=GRID, range=[0, y_top * 1.15])
    else:
        fig1 = go.Figure()
        _theme(fig1, "Рисковые виды спорта (ОСФ)", "нет рисковых зон в данных модели")
        fig1.update_layout(height=300)

    # ── ГРАФИК 3: pie по квадрантам ──
    quad_counts = df_all["Квадрант"].value_counts().reset_index()
    quad_counts.columns = ["Квадрант", "Кол-во"]
    fig3 = go.Figure(go.Pie(labels=quad_counts["Квадрант"], values=quad_counts["Кол-во"],
                            marker_colors=[QUADRANT_COLORS.get(q, "#94A3B8") for q in quad_counts["Квадрант"]],
                            textinfo="label+percent+value", hole=0.4, textfont=dict(size=11),
                            hovertemplate="<b>%{label}</b><br>Федераций: %{value}<br>%{percent}<extra></extra>"))
    _theme(fig3, "Распределение ОСФ по квадрантам", f"всего федераций: {len(df_all)}")
    fig3.update_layout(height=480, showlegend=False)

    parts = ["""<!DOCTYPE html><html lang="ru"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Антидопинговый дашборд: Риск × Рейтинг ОСФ 2025</title>
<style>
  body { font-family: Inter, Segoe UI, Roboto, Arial, sans-serif; background:#F8FAFC; color:#0F2D52; margin:0; }
  .header { background:#0F2D52; color:white; padding:24px 32px; }
  .header h1 { margin:0; font-size:22px; }
  .header p { margin:6px 0 0; font-size:13px; color:#7DD3FC; }
  .section { padding:24px 32px; }
  .section h2 { font-size:16px; border-left:4px solid #2563EB; padding-left:10px; margin-bottom:4px; }
  .section p { font-size:12px; color:#7C8DA6; margin:0 0 12px; }
  .divider { border:none; border-top:1px solid #EEF2F7; margin:0 32px; }
</style></head><body>
<div class="header">
  <h1>Антидопинговый дашборд: Матрица Риск × Рейтинг ОСФ 2025</h1>
  <p>Для работников антидопинговой сферы · Источники: Рейтинг ОСФ РУСАДА 2025 + прогнозная модель нарушений</p>
</div>"""]
    parts.append('<div class="section"><h2>Матрица Риск × Рейтинг</h2>'
                 '<p>Каждая точка — ОСФ. ⚪ серые точки внизу — федерации без данных модели. Наведите для деталей.</p>')
    # div_id фиксирован (не случайный uuid4 по умолчанию у Plotly) — иначе
    # идентичный прогон давал бы неидентичный HTML байт-в-байт (LOGIC.md §1).
    parts.append(fig2.to_html(full_html=False, include_plotlyjs="cdn", div_id="fig-osf-quadrant"))
    parts.append("</div><hr class='divider'>")
    parts.append('<div class="section"><h2>Топ рисковых видов спорта</h2>'
                 '<p>Только зоны 🔴 и 🟠 · сортировка по оценке модели</p>')
    parts.append(fig1.to_html(full_html=False, include_plotlyjs=False, div_id="fig-osf-top-risky"))
    parts.append("</div><hr class='divider'>")
    parts.append('<div class="section"><h2>Распределение по квадрантам</h2>'
                 '<p>Доля ОСФ в каждом квадранте матрицы</p>')
    parts.append(fig3.to_html(full_html=False, include_plotlyjs=False, div_id="fig-osf-pie"))
    parts.append("</div></body></html>")
    Path(path).write_text("\n".join(parts), encoding="utf-8")
    print(f"✅ HTML-дашборд сохранён: {path}")


# ══════════════════════════ ЗАПУСК ══════════════════════════

def main():
    ap = argparse.ArgumentParser(description="Анализ рисков ОСФ: рейтинг РУСАДА × модель")
    ap.add_argument("--pdf", default=str(DEFAULT_PDF))
    ap.add_argument("--risks-csv", default=str(DEFAULT_CSV))
    ap.add_argument("--ipynb", default=str(DEFAULT_IPYNB))
    ap.add_argument("--out", default=str(DEFAULT_OUT))
    args = ap.parse_args()

    pdf_path = Path(args.pdf)
    if not pdf_path.exists():
        print(f"❌ PDF рейтинга не найден: {pdf_path}\n"
              f"   (наблюдатель кладёт его в data/ratings/osf/latest.pdf)")
        sys.exit(2)

    print("📂 Загружаем рейтинг ОСФ из PDF...")
    osf_df = load_osf_rating(pdf_path)

    csv_path, ipynb_path = Path(args.risks_csv), Path(args.ipynb)
    if csv_path.exists():
        print(f"📈 Риски модели: CSV {csv_path} (полный вывод)")
        risks_df = load_risks_from_csv(csv_path)
    elif ipynb_path.exists():
        print(f"⚠ CSV модели нет — fallback на графики ноутбука {ipynb_path}.\n"
              f"  ВНИМАНИЕ: это только топ-N видов спорта; добавьте экспорт CSV в прогон модели.")
        risks_df = load_risks_from_ipynb(ipynb_path)
    else:
        print("⚠ Источники рисков не найдены — все ОСФ получат статус «нет данных модели»")
        risks_df = pd.DataFrame()
    print(f"   видов спорта в модели: {len(risks_df)}")

    print("🔗 Матчинг и сборка таблицы по всем ОСФ...")
    df_final, audit_df, unmatched_df = build_final_table(osf_df, risks_df)
    for z in [ZONE_CODE_RED, ZONE_CODE_ORANGE, ZONE_CODE_GREEN, ZONE_CODE_NODATA]:
        print(f"   {ZONE_EMOJI[z]} {z}: {(df_final['_zone_code'] == z).sum()}")

    out_dir = Path(args.out)
    out_dir.mkdir(parents=True, exist_ok=True)
    save_excel(df_final, audit_df, unmatched_df, out_dir / "osf_risk_final.xlsx")
    build_dashboard(df_final, out_dir / "osf_risk_dashboard_final.html")
    print(f"\n✅ Готово! Файлы в: {out_dir}")


if __name__ == "__main__":
    main()
