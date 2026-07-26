#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
siar/region_report.py — Анализ рисков субъектов РФ: зона модели × рейтинг регионов РУСАДА.
Двухосевая логика (утверждена): зона — ТОЛЬКО из модели (region_risks.csv, наихудший
сигнал по региону); балл рейтинга (порог 130 из 190) — вторая ось. Формулировки —
SIAR v2 из LOGIC.md §3 (без упоминания тестирования).

Запуск:
  python region_report.py --xlsx "Рейтинг регионов Итоги 2025.xlsx" \
                          --risks-csv region_risks.csv --out reports/
"""

import argparse
import re
import sys
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Правила приоритизации живут только в siar/rules.py (STRUCTURE.md, п. 3) —
# этот модуль их больше не дублирует, только вызывает evaluate().
ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))
from siar.rules import EntityKind, RATING_SCALES, evaluate  # noqa: E402
from siar.xlsx_determinism import normalize_xlsx_timestamps  # noqa: E402

try:
    import plotly.graph_objects as go
    import plotly.express as px
    HAS_PLOTLY = True
except ImportError:
    HAS_PLOTLY = False

# Порог не хардкодим — берём из siar/rules.py (единственное место правил).
RATING_THRESHOLD = RATING_SCALES[EntityKind.REGION].high_threshold
FO_SET = {"ЦФО", "СЗФО", "ЮФО", "СКФО", "ПФО", "УФО", "СФО", "ДФО", "НР"}

ZONE_ORDER = {"RED": 0, "ORANGE": 1, "GREEN": 2, "NO_DATA": 3}
ZONE_EMOJI = {"RED": "🔴", "ORANGE": "🟠", "GREEN": "🟢", "NO_DATA": "⚪"}
ZONE_FILLS = {"RED": "FDE2E2", "ORANGE": "FDF0D5", "GREEN": "DEF7EC", "NO_DATA": "F1F5F9"}
INK, SUB, GRID = "#0F2D52", "#7C8DA6", "#EEF2F7"
FONT = "Inter, Segoe UI, Roboto, Arial, sans-serif"
ZONE_COLORS = {"RED": "#DC2626", "ORANGE": "#F59E0B", "GREEN": "#10B981", "NO_DATA": "#94A3B8"}
QUADRANT_COLORS = {
    "Приоритет 1 (низкий рейтинг + систематичность)":      "#DC2626",
    "Приоритет 2 (высокий рейтинг + систематичность)":     "#F59E0B",
    "Приоритет 3 (низкий рейтинг + нет систематичности)":  "#3B82F6",
    "Приоритет 4 (высокий рейтинг + нет систематичности)": "#10B981",
}

# Колонка xlsx → (короткое имя критерия, тип): base — обычный, bonus/penalty — не входят
# в «невыполненные». Разметка по факту файла «Рейтинг регионов Итоги 2025».
CRITERIA_COLS = {
    2:  ("1.1 Ответственные/сайт", "base"),
    3:  ("1.2 Повышение квалификации", "base"),
    4:  ("1.3 Соглашение с РУСАДА", "base"),
    5:  ("1.4 План-график", "base"),
    6:  ("1.5 Отчёт за год", "base"),
    7:  ("1.6 Опрос", "base"),
    8:  ("1.7 Участие в мероприятиях", "base"),
    9:  ("1.8 Коммуникация", "base"),
    10: ("1.9 Мониторинг (штраф)", "penalty"),
    11: ("1.10 Организация мероприятий (бонус)", "bonus"),
    13: ("2.1 Квалификация лекторов", "base"),
    14: ("2.2 Обновление программ подготовки", "base"),
    15: ("2.3 НМС для спортшкол", "base"),
    16: ("2.4 Соглашение о перс. данных", "base"),
    17: ("2.5 Образование сборных региона", "base"),
    18: ("2.6 Стенды", "base"),
    22: ("3.1 Ответственный от Минздрава", "base"),
    23: ("3.2 Программа для медперсонала", "base"),
    24: ("3.3 Межведомственное взаимодействие", "base"),
    25: ("3.4 Инновационные системы (бонус)", "bonus"),
}
COL_B1, COL_B2, COL_B3, COL_TOTAL, COL_PLACE = 12, 19, 26, 27, 28

# Утверждённые формулировки LOGIC.md §3 теперь только в siar/rules.py —
# rules.evaluate() их и отдаёт (раньше здесь был второй, слегка разошедшийся
# набор REC_P1..REC_P4/JUSTIFICATION_*; например «совместно с регионом» вместо
# утверждённого «совместно с федерацией/регионом» — то самое расхождение).

QUADRANTS = {
    1: "Приоритет 1 (низкий рейтинг + систематичность)",
    2: "Приоритет 2 (высокий рейтинг + систематичность)",
    3: "Приоритет 3 (низкий рейтинг + нет систематичности)",
    4: "Приоритет 4 (высокий рейтинг + нет систематичности)",
}


def _norm_region(s):
    s = str(s).lower().replace("ё", "е")
    s = re.sub(r"\([^)]*\)", " ", s)
    s = re.sub(r"[\-–—−]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


REGION_ALIASES = {  # варианты написания → канонический ключ (применяется к обеим сторонам)
    "кемеровская кузбасс область": "кемеровская область",
    "республика чувашская": "республика чувашия",
    "республика удмуртская": "республика удмуртия",
    "карачаево республика черкесская": "карачаево республика черкесия",
    "балкарская кабардино республика": "балкария кабардино республика",
    "ао ненецкий ямало": "автономный ненецкий округ ямало",
    "ао мансийский ханты": "автономный мансийский округ ханты югра",
    "автономный мансийский округ ханты": "автономный мансийский округ ханты югра",
    "хмао югра": "автономный мансийский округ ханты югра",
}


def _key(s):
    toks = sorted(_norm_region(s).split())
    k = " ".join(toks)
    return REGION_ALIASES.get(k, k)


# ══════════════════ ЗАГРУЗКА РЕЙТИНГА ══════════════════

def load_rating(xlsx_path):
    raw = pd.read_excel(xlsx_path, header=None)
    rows, dropped_headers = [], 0
    for _, r in raw.iterrows():
        fo, name = str(r[0]).strip(), str(r[1]).strip()
        if fo not in FO_SET:
            continue
        if name in ("", "nan", "Субъект РФ"):
            dropped_headers += 1
            continue
        crit = {}
        for col, (label, kind) in CRITERIA_COLS.items():
            v = pd.to_numeric(r[col], errors="coerce")
            crit[label] = (0.0 if pd.isna(v) else float(v), kind)
        def _num(col):
            v = pd.to_numeric(r[col], errors="coerce")
            return 0.0 if pd.isna(v) else float(v)
        rows.append({
            "ФО": fo, "Регион": re.sub(r"\s+", " ", name),
            "_crit": crit,
            "Блок 1": _num(COL_B1), "Блок 2": _num(COL_B2), "Блок 3": _num(COL_B3),
            "Итого баллов": _num(COL_TOTAL),
            "Место": int(_num(COL_PLACE)) if _num(COL_PLACE) else "",
        })
    df = pd.DataFrame(rows).drop_duplicates(subset=["Регион"]).reset_index(drop=True)
    df["_key"] = df["Регион"].apply(_key)

    def done(c):    return "; ".join(l for l, (v, k) in c.items() if k == "base" and v > 0) or "нет"
    def notdone(c): return "; ".join(l for l, (v, k) in c.items() if k == "base" and v <= 0) or "нет"
    def extras(c):
        out = []
        for l, (v, k) in c.items():
            if k == "bonus" and v > 0: out.append(f"{l}: +{v:g}")
            if k == "penalty" and v < 0: out.append(f"{l}: {v:g}")
        return "; ".join(out)
    df["Выполненные критерии"] = df["_crit"].apply(done)
    df["Невыполненные критерии"] = df["_crit"].apply(notdone)
    df["Бонусы/штрафы"] = df["_crit"].apply(extras)
    print(f"   строк-заголовков отфильтровано: {dropped_headers} · регионов: {len(df)}")
    return df


# ══════════════════ SIAR v2 ══════════════════

def build_table(rating_df, risks_df):
    risk_by_key, audit, unmatched = {}, [], []
    if not risks_df.empty:
        risks_df = risks_df.copy()
        risks_df["_key"] = risks_df["Субъект РФ"].apply(_key)
        rating_keys = set(rating_df["_key"])
        for _, rr in risks_df.iterrows():
            hit = rr["_key"] in rating_keys
            audit.append({"Регион (модель)": rr["Субъект РФ"],
                          "Ключ": rr["_key"], "Найден в рейтинге": "да" if hit else "нет",
                          "Зона": rr["зона_риска"], "proba": round(float(rr["proba"]), 3)})
            if hit:
                risk_by_key[rr["_key"]] = rr
            else:
                unmatched.append({"Название": rr["Субъект РФ"], "Источник": "модель",
                                  "Причина": "регион не найден в рейтинге РУСАДА"})

    out = []
    for _, r in rating_df.iterrows():
        score = float(r["Итого баллов"])
        miss, extras = r["Невыполненные критерии"], r["Бонусы/штрафы"]
        unmet = [] if miss == "нет" else [c.strip() for c in miss.split("; ") if c.strip()]
        rr = risk_by_key.get(r["_key"])
        zone = "NO_DATA" if rr is None else str(rr["зона_риска"])
        proba = None if rr is None else float(rr["proba"])
        reason = None if rr is None else (str(rr["причина"]) or None)

        # Приоритет/состояние/обоснование/рекомендация — только через rules.evaluate()
        # (LOGIC.md §3, дословные формулировки; никакой генерации «на лету» здесь).
        try:
            a = evaluate(EntityKind.REGION, zone, score, reason=reason,
                        unmet_criteria=unmet, object_name=r["Регион"])
        except ValueError as e:
            sys.exit(f"❌ SIAR v2 отклонил «{r['Регион']}»: {e}")

        out.append({
            "ФО": r["ФО"], "Регион": r["Регион"], "Приоритет": a.priority,
            "Квадрант": QUADRANTS[a.priority],
            "Зона риска": f"{ZONE_EMOJI[zone]} {zone}", "_zone_code": zone,
            "Статус": a.status or "",
            "Оценка риска (proba)": round(proba, 3) if proba is not None else None,
            "Итого баллов": score, "Место": r["Место"],
            "Блок 1": r["Блок 1"], "Блок 2": r["Блок 2"], "Блок 3": r["Блок 3"],
            "Выполненные критерии": r["Выполненные критерии"],
            "Невыполненные критерии": miss,
            "Бонусы/штрафы": extras,
            "Рекомендация": a.recommendation, "Обоснование": a.justification,
        })

    df = pd.DataFrame(out)
    df["_zs"] = df["_zone_code"].map(ZONE_ORDER)
    df["_p"] = df["Оценка риска (proba)"].fillna(-1)
    df = (df.sort_values(["Приоритет", "_zs", "_p", "Итого баллов"],
                         ascending=[True, True, False, False])
            .drop(columns=["_zs", "_p"]).reset_index(drop=True))
    df.insert(0, "№", range(1, len(df) + 1))

    # ── Инварианты ──
    ok = True
    n = len(rating_df)
    print(f"\n{'='*64}\nСАМОПРОВЕРКА (инварианты)\n{'='*64}")
    c1 = len(df) == n; ok &= c1
    print(f"1. Строк в итоге ({len(df)}) = N регионов ({n}): {'✓' if c1 else '✗'}")
    pr = df["Приоритет"].value_counts().sort_index().to_dict()
    c2 = sum(pr.values()) == len(df); ok &= c2
    print(f"2. Приоритеты {pr} · сумма {sum(pr.values())}: {'✓' if c2 else '✗'}")
    print(f"3. Зоны: {df['_zone_code'].value_counts().to_dict()}")
    b4 = df[df["_zone_code"].isin(["RED","ORANGE"]) & df["Приоритет"].isin([3,4])]
    b5 = df[(df["_zone_code"]=="GREEN") & df["Приоритет"].isin([1,2])]
    c45 = len(b4)==0 and len(b5)==0; ok &= c45
    print(f"4-5. Зоны↔приоритеты нарушений: {len(b4)+len(b5)}: {'✓' if c45 else '✗'}")
    hi = df["Итого баллов"] >= RATING_THRESHOLD
    b6 = df[(hi & df["Приоритет"].isin([1,3])) | (~hi & df["Приоритет"].isin([2,4]))]
    c6 = len(b6)==0; ok &= c6
    print(f"6. Порог {RATING_THRESHOLD}↔приоритеты нарушений: {len(b6)}: {'✓' if c6 else '✗'}")
    b7 = df[df["Регион"].str.strip() == "Субъект РФ"]
    c7 = len(b7)==0; ok &= c7
    print(f"7. Строк-заголовков в данных: {len(b7)}: {'✓' if c7 else '✗'}")
    print("Средний балл по ФО:",
          df.groupby("ФО")["Итого баллов"].mean().round(0).astype(int).sort_values(ascending=False).to_dict())
    print(f"{'='*64}: {'ВСЕ ИНВАРИАНТЫ ПРОЙДЕНЫ' if ok else 'НАРУШЕНИЯ — ВЫДАЧА ЗАПРЕЩЕНА'}\n")
    if not ok:
        sys.exit(3)

    fo = df.groupby("ФО").agg(
        Регионов=("Регион","count"),
        RED=("_zone_code", lambda s:(s=="RED").sum()),
        ORANGE=("_zone_code", lambda s:(s=="ORANGE").sum()),
        GREEN=("_zone_code", lambda s:(s=="GREEN").sum()),
        Нет_данных=("_zone_code", lambda s:(s=="NO_DATA").sum()),
        Средний_балл=("Итого баллов","mean"),
        Мин=("Итого баллов","min"), Макс=("Итого баллов","max"),
    ).round(1).reset_index().sort_values("Средний_балл", ascending=False)

    return df, fo, pd.DataFrame(audit), pd.DataFrame(unmatched, columns=["Название","Источник","Причина"])


# ══════════════════ EXCEL ══════════════════

def save_excel(df, fo, audit, unmatched, path):
    export = df[[c for c in df.columns if not c.startswith("_")]]
    zones = df["_zone_code"].tolist()
    with pd.ExcelWriter(path, engine="openpyxl") as xl:
        export.to_excel(xl, index=False, sheet_name="Матрица регионов")
        fo.to_excel(xl, index=False, sheet_name="Сводка по ФО")
        unmatched.to_excel(xl, index=False, sheet_name="Не сопоставлено")
        if not audit.empty:
            audit.to_excel(xl, index=False, sheet_name="Аудит матчинга")
    wb = openpyxl.load_workbook(path)
    ws = wb["Матрица регионов"]
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="0F2D52")
        cell.font = Font(color="FFFFFF", bold=True, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    widths = {"№":5,"ФО":7,"Регион":30,"Приоритет":10,"Квадрант":30,"Зона риска":13,
              "Статус":18,"Оценка риска (proba)":12,"Итого баллов":10,"Место":8,
              "Блок 1":8,"Блок 2":8,"Блок 3":8,"Выполненные критерии":40,
              "Невыполненные критерии":40,"Бонусы/штрафы":30,
              "Рекомендация":50,"Обоснование":62}
    for i, name in enumerate(export.columns, 1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(name, 16)
    thin = Side(style="thin", color="D1D5DB"); border = Border(thin, thin, thin, thin)
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row)):
        fill = PatternFill("solid", fgColor=ZONE_FILLS.get(zones[i], "FFFFFF"))
        for c in row:
            c.alignment = Alignment(vertical="top", wrap_text=True); c.border = border; c.fill = fill
    ws.row_dimensions[1].height = 42
    for i in range(2, ws.max_row + 1):
        ws.row_dimensions[i].height = 46
    ws.freeze_panes = "C2"
    # Воспроизводимость (LOGIC.md §1): фиксируем created/modified — иначе
    # openpyxl пишет текущее время в docProps/core.xml, и идентичный вход
    # давал бы неидентичный файл байт-в-байт (данные это не меняет).
    _FIXED_TIMESTAMP = datetime(2000, 1, 1)
    wb.properties.created = _FIXED_TIMESTAMP
    wb.properties.modified = _FIXED_TIMESTAMP
    wb.save(path)
    normalize_xlsx_timestamps(path)  # zip-контейнер: см. siar/xlsx_determinism.py
    print(f"✅ Excel сохранён: {path}")



def _theme(fig, title, subtitle=""):
    sub_html = f"<br><span style=\'font-size:12px;color:{SUB}\'>{subtitle}</span>" if subtitle else ""
    fig.update_layout(template="plotly_white", font=dict(family=FONT, color=INK, size=13),
                      title=dict(text=f"<b>{title}</b>{sub_html}", x=0.01, xanchor="left", y=0.97,
                                 font=dict(size=18)),
                      margin=dict(l=10, r=24, t=84, b=24),
                      paper_bgcolor="white", plot_bgcolor="white", legend_title_text="")
    fig.update_xaxes(showgrid=False, zeroline=False)
    fig.update_yaxes(showgrid=False, zeroline=False)
    return fig


def build_dashboard(df, fo, path):
    if not HAS_PLOTLY:
        print("\u26a0 plotly не установлен — дашборд регионов пропущен (pip install plotly). Excel готов.")
        return
    df_all = df.copy()
    p_max = float(pd.to_numeric(df_all["Оценка риска (proba)"], errors="coerce").max() or 0)
    y_top = max(0.1, p_max * 1.25)

    # ── 1. Матрица квадрантов (scatter) ──
    fig1 = go.Figure()
    rng = np.random.default_rng(42)
    for quad in QUADRANT_COLORS:
        sub = df_all[df_all["Квадрант"] == quad].copy()
        if sub.empty:
            continue
        y = pd.to_numeric(sub["Оценка риска (proba)"], errors="coerce").fillna(-y_top * 0.07).values
        x = pd.to_numeric(sub["Итого баллов"], errors="coerce").fillna(0).values.astype(float)
        x = x + rng.uniform(-3, 3, size=len(x))
        y = y + rng.uniform(-y_top * 0.015, y_top * 0.015, size=len(y))
        fig1.add_trace(go.Scatter(
            x=x, y=y, mode="markers", name=quad,
            marker=dict(color=QUADRANT_COLORS[quad], size=11, opacity=0.82,
                        line=dict(width=1, color="white")),
            customdata=np.stack([sub["Регион"].values, sub["ФО"].values,
                                 sub["Зона риска"].values, sub["Обоснование"].values], axis=-1),
            hovertemplate=("<b>%{customdata[0]}</b> (%{customdata[1]})<br>Баллы: %{x:.0f} из 190<br>"
                           "Зона: %{customdata[2]}<br><i>%{customdata[3]}</i><extra></extra>")))
    fig1.add_vline(x=RATING_THRESHOLD, line_dash="dot", line_color="#94A3B8", line_width=1.5,
                   annotation_text=f"порог {RATING_THRESHOLD}", annotation_position="top left",
                   annotation_font=dict(color="#94A3B8", size=10))
    fig1.add_hline(y=0.0, line_dash="dot", line_color="#CBD5E1", line_width=1)
    _theme(fig1, "Матрица Риск × Рейтинг регионов",
           "ось X — баллы рейтинга (из 190) · ось Y — оценка риска модели (⚪ нет данных ≈ ниже 0)")
    fig1.update_layout(height=640,
        xaxis=dict(range=[0, 200], title="Баллы рейтинга РУСАДА", showgrid=True, gridcolor=GRID),
        yaxis=dict(range=[-y_top * 0.14, y_top], title="Оценка риска (proba)", showgrid=True, gridcolor=GRID),
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1))

    # ── 2. Топ рисковых регионов (RED+ORANGE) ──
    risk = df_all[df_all["_zone_code"].isin(["RED", "ORANGE"])].copy()
    if not risk.empty:
        risk["_v"] = pd.to_numeric(risk["Итого баллов"], errors="coerce")
        d = risk.sort_values("_v").tail(20)
        fig2 = px.bar(d, x="Итого баллов", y="Регион", orientation="h",
                      color="_zone_code", color_discrete_map=ZONE_COLORS,
                      text=d["Итого баллов"].map(lambda v: f"{v:g}"),
                      hover_data={"ФО": True, "Квадрант": True, "_zone_code": False})
        fig2.update_traces(marker_cornerradius=5, textposition="outside", cliponaxis=False,
                           textfont=dict(size=11, color=INK))
        _theme(fig2, "Регионы с повышенным риском", "зоны 🔴 и 🟠 · топ-20 по баллам рейтинга")
        fig2.update_layout(height=max(420, 30 * len(d)), yaxis_title="", xaxis_title="Итого баллов (из 190)",
                           legend_title_text="зона")
        fig2.update_xaxes(showgrid=True, gridcolor=GRID)
    else:
        fig2 = go.Figure(); _theme(fig2, "Регионы с повышенным риском", "нет рисковых зон"); fig2.update_layout(height=300)

    # ── 3. Распределение по приоритетам (pie) ──
    pc = df_all["Квадрант"].value_counts().reset_index()
    pc.columns = ["Квадрант", "n"]
    fig3 = go.Figure(go.Pie(labels=pc["Квадрант"], values=pc["n"],
                            marker_colors=[QUADRANT_COLORS.get(q, "#94A3B8") for q in pc["Квадрант"]],
                            textinfo="label+percent+value", hole=0.4, textfont=dict(size=11),
                            hovertemplate="<b>%{label}</b><br>Регионов: %{value}<br>%{percent}<extra></extra>"))
    _theme(fig3, "Распределение регионов по квадрантам", f"всего субъектов: {len(df_all)}")
    fig3.update_layout(height=460, showlegend=False)

    # ── 4. Средний балл по ФО ──
    fob = fo.sort_values("Средний_балл")
    fig4 = px.bar(fob, x="Средний_балл", y="ФО", orientation="h",
                  text=fob["Средний_балл"].map(lambda v: f"{v:g}"))
    clr = ["#10B981" if v >= RATING_THRESHOLD else ("#F59E0B" if v >= 95 else "#DC2626") for v in fob["Средний_балл"]]
    fig4.update_traces(marker_color=clr, marker_cornerradius=5, textposition="outside", cliponaxis=False,
                       textfont=dict(size=11, color=INK))
    fig4.add_vline(x=RATING_THRESHOLD, line_dash="dot", line_color="#94A3B8", line_width=1.5)
    _theme(fig4, "Средний балл по федеральным округам", "цвет — уровень среднего балла ФО")
    fig4.update_layout(height=420, yaxis_title="", xaxis_title="Средний балл (из 190)")
    fig4.update_xaxes(showgrid=True, gridcolor=GRID, range=[0, 200])

    blocks = [
        ("Матрица Риск × Рейтинг", "Каждая точка — субъект РФ. ⚪ серые внизу — регионы без данных модели. Наведите для деталей.", fig1, "cdn"),
        ("Топ рисковых регионов", "Только зоны 🔴 и 🟠 · сортировка по баллам", fig2, False),
        ("Распределение по квадрантам", "Доля субъектов в каждом квадранте матрицы", fig3, False),
        ("Средний балл по ФО", "Сравнение федеральных округов", fig4, False),
    ]
    parts = ["""<!DOCTYPE html><html lang="ru"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Антидопинговый дашборд: Риск × Рейтинг регионов 2025</title>
<style>
  body { font-family: Inter, Segoe UI, Roboto, Arial, sans-serif; background:#F8FAFC; color:#0F2D52; margin:0; }
  .header { background:#0F2D52; color:white; padding:24px 32px; }
  .header h1 { margin:0; font-size:22px; } .header p { margin:6px 0 0; font-size:13px; color:#7DD3FC; }
  .section { padding:24px 32px; } .section h2 { font-size:16px; border-left:4px solid #2563EB; padding-left:10px; margin-bottom:4px; }
  .section p { font-size:12px; color:#7C8DA6; margin:0 0 12px; } .divider { border:none; border-top:1px solid #EEF2F7; margin:0 32px; }
</style></head><body>
<div class="header"><h1>Антидопинговый дашборд: Матрица Риск × Рейтинг регионов 2025</h1>
<p>Для работников антидопинговой сферы · Источники: Рейтинг регионов РУСАДА 2025 + прогнозная модель нарушений</p></div>"""]
    # div_id фиксирован по порядковому номеру (не случайный uuid4 по умолчанию
    # у Plotly) — иначе идентичный прогон давал бы неидентичный HTML
    # байт-в-байт (LOGIC.md §1).
    for idx, (title, sub, fig, js) in enumerate(blocks):
        parts.append(f'<div class="section"><h2>{title}</h2><p>{sub}</p>')
        parts.append(fig.to_html(full_html=False, include_plotlyjs=js, div_id=f"fig-region-{idx}"))
        parts.append("</div><hr class=\'divider\'>")
    parts.append("</body></html>")
    Path(path).write_text("\n".join(parts), encoding="utf-8")
    print(f"✅ HTML-дашборд сохранён: {path}")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True)
    ap.add_argument("--risks-csv", required=True)
    ap.add_argument("--out", default=".")
    a = ap.parse_args()
    Path(a.out).mkdir(parents=True, exist_ok=True)
    print("📂 Загружаем рейтинг регионов...")
    rating = load_rating(a.xlsx)
    risks = pd.read_csv(a.risks_csv)
    print(f"   регионов в модели: {len(risks)}")
    df, fo, audit, unmatched = build_table(rating, risks)
    for z in ["RED","ORANGE","GREEN","NO_DATA"]:
        print(f"   {ZONE_EMOJI[z]} {z}: {(df['_zone_code']==z).sum()}")
    save_excel(df, fo, audit, unmatched, Path(a.out) / "region_risk_final.xlsx")
    build_dashboard(df, fo, Path(a.out) / "region_risk_dashboard.html")


if __name__ == "__main__":
    main()
